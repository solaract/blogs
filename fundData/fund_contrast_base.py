import os
import akshare as ak
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Alignment
from datetime import datetime
import requests
from bs4 import BeautifulSoup
import json
import re
from pathlib import Path
from tqdm import tqdm

# ---------- 新增关键设置 ----------
# 获取当前脚本所在目录
script_dir = Path(__file__).parent.absolute()
os.chdir(script_dir)  # 设置工作目录

# 确保目标文件存在
if not Path('沪深300被动对比.xlsx').exists():
    wb = Workbook()
    wb.save('沪深300被动对比.xlsx')
# -------------------------------

# 1. 初始化主表格
wb = load_workbook('沪深300被动对比.xlsx')
ws = wb.active
ws['A1'] = datetime.now().strftime('%Y-%m-%d')

# 2. 检查代码文件存在
if not Path('沪深300被动代码.xlsx').exists():
    raise FileNotFoundError("请先创建 沪深300被动代码.xlsx 文件并在A1开始按列填写基金代码")

# 3. 读取基金代码

# 强制指定列为字符串类型
code_df = pd.read_excel(
    '沪深300被动代码.xlsx',
    header=None,
    dtype={0: str}  # 假设第0列需要保留前导零
)
fund_codes = code_df.iloc[:, 0].dropna().tolist()
# print(fund_codes)

# 3. 准备数据存储结构
output_list = []
output_data = {
    '基金代码': '',
    '基金名称': '',
    '成立日期': '',
    '今年来': '', '近1周': '', '近1月': '', '近3月': '', 
    '近6月': '', '近1年': '', '近2年': '', '近3年': '','近5年':'',
    '年化跟踪误差': '',
    '管理费率': '',
    '托管费率': '',
    '基金规模': ''
}

# # 4. 单独处理标的指数收益率（B列）
# yield_type_map = {
#     'sy': ('今年', 4),
#     'm': ('1月', 6),
#     'q': ('3月', 7),
#     'hy': ('6月', 8),
#     'y': ('1年', 9),
#     'try': ('3年', 11),
#     'fiy':('5年',12)
# }

# # 获取沪深300指数数据
# # hs300_data = {}
# for t, (name, row) in yield_type_map.items():
#     # print(t)
#     url = f'https://api.fund.eastmoney.com/pinzhong/LJSYLZS?fundCode=110020&indexcode=000300&type={t}&callback=yield'
#     response = requests.get(url, headers={'Referer': 'https://fund.eastmoney.com/110020.html?spm=search',
#                                           'Host': 'api.fund.eastmoney.com'})
#     # print(url)
#     # print(response.text)
#     json_str = response.text[6:-1]
#     data = json.loads(json_str)
#     # hs300_item = next(item for item in data['Data'] if item['name'] == "沪深300")
#     # hs300_data[name] = data['Data'][2]['data'][-1][1]
#     hs300_data = data['Data'][2]['data'][-1][1]
#     ws[f'B{row}'] = f'{hs300_data}%'
#     # print(hs300_data)

# 写入B列
# ws['B4'] = hs300_data.get('今年', '')
# ws['B6'] = hs300_data.get('1月', '')
# ws['B7'] = hs300_data.get('3月', '')
# ws['B8'] = hs300_data.get('6月', '')
# ws['B9'] = hs300_data.get('1年', '')
# ws['B11'] = hs300_data.get('3年', '')
# ws['B12'] = hs300_data.get('5年', '')

# 5. 处理各基金数据（使用pandas批量处理）
fund_info_df = ak.fund_info_index_em(symbol="沪深指数", indicator="被动指数型")

color = ['F8CBAD','FFC000','8EA9DB','A9D08E','D9D9D9']
alig = Alignment('center','center')
sum1 = 0
sum2 = 0
sum3 = 0
sum5 = 0


# 创建进度条
with tqdm(fund_codes, 
    desc='获取基金数据', 
    unit='支',
    ncols=100) as pbar:
    for code in pbar: 
        try:
            pbar.set_postfix_str(f"当前处理：{code}")
            # 获取基本信息
            fund_info = fund_info_df[fund_info_df['基金代码'] == code].iloc[0].to_dict()

            # print(fund_info)

            # 填充数据字典
            output_data['基金代码'] = code
            output_data['基金名称'] = fund_info['基金名称']
            
            # 网页数据抓取
            res = requests.get(f'https://fund.eastmoney.com/{code}.html')
            res.encoding = "utf-8"
            # print(res.text)
            soup = BeautifulSoup(res.text, 'html.parser')
            # print(soup)
            
            # 成立日期、年化跟踪误差、基金规模
            if soup:
                # span_tag = soup.find_all('div', string=re.compile(r'成\s*立\s*日'))
                fund_info_tag = soup.find('div', class_='infoOfFund')
                date_tag = fund_info_tag.table.find_all('tr')[1].td
                # print(date_tag)
                # print(repr(date_tag))
                # for string in date_tag.strings:
                #      print(repr(string))
                # a = soup.find('a', string='年化跟踪误差：')
                deviation_tag = fund_info_tag.find('td', class_='specialData')
                # print(deviation_tag.text)
                sclae_tag = fund_info_tag.table.tr.contents[1]
                # print(sclae_tag.text)
                # print(date_tag.text)
                date_data = re.search(r'\d{4}-\d{2}-\d{2}',date_tag.text)
                deviation_data = re.search(r'\d+\.\d+%', deviation_tag.text)
                sclae_data = re.search(r'(\d+\.\d+)亿元', sclae_tag.text)
                output_data['成立日期'] = date_data.group() if date_data else ''
                # print(output_data)
                output_data['年化跟踪误差'] = deviation_data.group() if deviation_data else ''
                output_data['基金规模'] = sclae_data.group(1) if sclae_data else ''
                # print(output_data)
            
            # # 填充收益率指标数据
            # for key in ['今年来', '近1周', '近1月', '近3月', '近6月', '近1年', '近2年', '近3年']:
            #     if fund_info[key]:
            #          output_data[key] = f'{fund_info[key]}%'
            #          match key:
            #             case '近1年':
            #                 sum1 = sum1 + fund_info[key]
            #             case '近2年':
            #                 sum2 = sum2 + fund_info[key]
            #             case '近3年':
            #                 sum3 = sum3 + fund_info[key]
            #     else:
            #          output_data[key] = ''

            
            # 获取5年收益率
            # url = f'https://api.fund.eastmoney.com/pinzhong/LJSYLZS?fundCode={code}&indexcode=000300&type=fiy&callback=yield'
            # response = requests.get(url, headers={'Referer': f'https://fund.eastmoney.com/{code}.html?spm=search',
            #                                     'Host': 'api.fund.eastmoney.com'})
            # json_str = response.text[6:-1]
            # data = json.loads(json_str)
            # fiy_data = data['Data'][0]['data'][-1][1]
            # output_data['近5年'] = f'{fiy_data}%'
            # sum5 = sum5 + fiy_data

            # 获取收益率
            rise_url = f'https://fundf10.eastmoney.com/FundArchivesDatas.aspx?type=jdzf&code={code}'
            rise_res = requests.get(rise_url)
            # print(rise_res.text)
            rise_html = re.findall(r"content:\s*\"(.*?)\"\s*};", rise_res.text, re.DOTALL)[0]
            # rise_res.encoding = "utf-8"
            rise_soup = BeautifulSoup(rise_html, 'html.parser')
            rise_data = rise_soup.div.contents

            if sum1 == 0:
                for i in range(1,10):
                    hs300_data = rise_data[i].contents[3].text
                    ws[f'B{i+3}'] = hs300_data
                    ws[f'B{i+3}'].alignment = alig


            output_data['今年来'] = rise_data[1].contents[1].text
            output_data['近1周'] = rise_data[2].contents[1].text
            output_data['近1月'] = rise_data[3].contents[1].text
            output_data['近3月'] = rise_data[4].contents[1].text
            output_data['近6月'] = rise_data[5].contents[1].text
            output_data['近1年'] = rise_data[6].contents[1].text
            sum1 = sum1 + float(output_data['近1年'].split('%')[0])
            output_data['近2年'] = rise_data[7].contents[1].text
            sum2 = sum2 + float(output_data['近2年'].split('%')[0])
            output_data['近3年'] = rise_data[8].contents[1].text
            sum3 = sum3 + float(output_data['近3年'].split('%')[0])
            output_data['近5年'] = rise_data[9].contents[1].text
            sum5 = sum5 + float(output_data['近5年'].split('%')[0])
            
            
            

            # print(output_data)
            
            # 费率信息
            fee_res = requests.get(f'https://fundf10.eastmoney.com/jbgk_{code}.html')
            fee_res.encoding = "utf-8"
            fee_soup = BeautifulSoup(fee_res.text, 'html.parser')
            manage_data = fee_soup.find('th', string='管理费率').find_next('td').text
            custodian_data = fee_soup.find('th', string='托管费率').find_next('td').text
            output_data['管理费率'] = manage_data.split('%')[0] + '%' if manage_data else ''
            output_data['托管费率'] = custodian_data.split('%')[0] + '%' if custodian_data else ''
            
            output_list.append(output_data.copy())
            # print(output_data)
            # break




        except Exception as e:
            pbar.write(f"⚠️ 基金 {code} 数据获取失败: {str(e)}")
            continue

# 6. 将处理好的数据批量写入Excel（从C列开始）
# print(output_data)
# result_df = pd.DataFrame(output_data)
# for idx, row in result_df.iterrows():
#     col = chr(67 + idx)  # C列开始
#     ws[f'{col}1'] = row['基金代码']
#     ws[f'{col}2'] = row['基金名称']
#     ws[f'{col}3'] = row['成立日期']
    
#     # 写入指标数据
#     for i, key in enumerate(['今年来', '近1周', '近1月', '近3月', '近6月', '近1年', '近2年', '近3年'], start=4):
#         ws[f'{col}{i}'] = row[key]
    
#     # 写入其他数据
#     ws[f'{col}13'] = row['年化跟踪误差']
#     ws[f'{col}14'] = row['管理费率']
#     ws[f'{col}15'] = row['托管费率']
#     ws[f'{col}16'] = row['基金规模']

# print(output_list)

len = len(output_list)
ave1 = sum1 / len
ave2 = sum2 / len
ave3 = sum3 / len
ave5 = sum5 / len
print(ave5)
ave1_tag = 0
ave2_tag = 0
ave3_tag = 0
ave5_tag = 0
for i in range(len):
    col = chr(67 + i)
    ws[f'{col}1'] = output_list[i]['基金代码']
    ws[f'{col}1'].alignment = alig
    ws[f'{col}2'] = output_list[i]['基金名称']
    ws[f'{col}2'].alignment = alig
    ws[f'{col}2'].fill = PatternFill(fill_type=None)
    ws[f'{col}3'] = output_list[i]['成立日期']
    ws[f'{col}3'].alignment = alig
# 写入指标数据
    for j, key in enumerate(['今年来', '近1周', '近1月', '近3月', '近6月', '近1年', '近2年', '近3年','近5年'], start=4):
        ws[f'{col}{j}'] = output_list[i][key]
        ws[f'{col}{j}'].alignment = alig
        ws[f'{col}{j}'].fill = PatternFill(fill_type=None)
        match key:
            case '近1年':
                if float(output_list[i][key].split('%')[0]) > ave1:
                    ws[f'{col}{j}'].fill = PatternFill('solid',fgColor=color[0])
                    ave1_tag = 1
            case '近2年':
                if float(output_list[i][key].split('%')[0]) > ave2:
                    ws[f'{col}{j}'].fill = PatternFill('solid',fgColor=color[0])
                    ave2_tag = 1
            case '近3年':
                if float(output_list[i][key].split('%')[0]) > ave3:
                    ws[f'{col}{j}'].fill = PatternFill('solid',fgColor=color[0])
                    ave3_tag = 1
            case '近5年':
                if float(output_list[i][key].split('%')[0]) > ave5:
                    ws[f'{col}{j}'].fill = PatternFill('solid',fgColor=color[0])
                    ave5_tag = 1
    
    # 写入其他数据
    ws[f'{col}13'] = output_list[i]['年化跟踪误差']
    ws[f'{col}13'].fill = PatternFill('solid',fgColor=color[1])
    ws[f'{col}13'].alignment = alig
    ws[f'{col}14'] = output_list[i]['管理费率']
    ws[f'{col}14'].fill = PatternFill('solid',fgColor=color[2])
    ws[f'{col}14'].alignment = alig
    ws[f'{col}15'] = output_list[i]['托管费率']
    ws[f'{col}15'].fill = PatternFill('solid',fgColor=color[2])
    ws[f'{col}15'].alignment = alig
    ws[f'{col}16'] = output_list[i]['基金规模']
    ws[f'{col}16'].alignment = alig
    ws[f'{col}16'].fill = PatternFill('solid',fgColor=color[4])

    if ave1_tag and ave2_tag and ave3_tag and ave5_tag:
        ws[f'{col}2'].fill = PatternFill('solid',fgColor=color[3])
    ave1_tag = 0
    ave2_tag = 0
    ave3_tag = 0
    ave5_tag = 0

     

# 保存文件
wb.save('沪深300被动对比.xlsx')
print("处理完成！")