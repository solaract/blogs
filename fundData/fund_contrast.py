import os
import re
from datetime import datetime
from pathlib import Path
from typing import Dict, List

import akshare as ak
import pandas as pd
import requests
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, PatternFill
from openpyxl.worksheet.worksheet import Worksheet
from tqdm import tqdm
from bs4 import BeautifulSoup

# ----------------- 常量配置 -----------------
BASE_DIR = Path(__file__).parent.absolute()
CONFIG = {
    "input_codes_file": "沪深300被动代码.xlsx",
    "output_file": "沪深300被动对比.xlsx",
    "colors": {
        "performance": "F8CBAD",
        "tracking_error": "FFC000",
        "fee": "8EA9DB",
        "highlight": "A9D08E",
        "scale": "D9D9D9"
    },
    "headers": [
        '基金代码', '基金名称', '成立日期', '今年来', '近1周', '近1月', '近3月',
        '近6月', '近1年', '近2年', '近3年', '近5年', '年化跟踪误差',
        '管理费率', '托管费率', '基金规模'
    ],
    "user_agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
}
# ------------------------------------------

def initialize_workbook() -> Workbook:
    """初始化或创建Excel工作簿"""
    output_path = BASE_DIR / CONFIG["output_file"]
    if not output_path.exists():
        wb = Workbook()
        wb.save(output_path)
    return load_workbook(output_path)

def read_fund_codes() -> List[str]:
    """读取基金代码文件"""
    codes_path = BASE_DIR / CONFIG["input_codes_file"]
    if not codes_path.exists():
        raise FileNotFoundError(f"请先创建 {CONFIG['input_codes_file']} 文件")

    code_df = pd.read_excel(
        codes_path,
        header=None,
        dtype={0: str}
    )
    return code_df.iloc[:, 0].dropna().tolist()

def safe_request(url: str, max_retries: int = 3) -> requests.Response:
    """带重试机制的请求函数"""
    headers = {"User-Agent": CONFIG["user_agent"]}
    for _ in range(max_retries):
        try:
            response = requests.get(url, headers=headers, timeout=10)
            response.raise_for_status()
            return response
        except requests.RequestException as e:
            tqdm.write(f"请求失败: {url}, 错误: {str(e)}")
    raise requests.RequestException(f"请求失败超过最大重试次数: {url}")

def parse_fund_basic_info(soup: BeautifulSoup) -> Dict[str, str]:
    """解析基金基本信息"""
    result = {
        "成立日期": "",
        "年化跟踪误差": "",
        "基金规模": ""
    }

    info_div = soup.find("div", class_="infoOfFund")
    if not info_div or not info_div.table:
        return result

    try:
        # 成立日期
        date_tag = info_div.table.find_all("tr")[1].td
        if date_tag:
            date_match = re.search(r"\d{4}-\d{2}-\d{2}", date_tag.text)
            result["成立日期"] = date_match.group() if date_match else ""
        
        # 跟踪误差
        error_tag = info_div.find("td", class_="specialData")
        if error_tag:
            error_match = re.search(r"\d+\.\d+%", error_tag.text)
            result["年化跟踪误差"] = error_match.group() if error_match else ""
        
        # 基金规模
        scale_tag = info_div.table.tr.contents[1]
        if scale_tag:
            scale_match = re.search(r"(\d+\.\d+)亿元", scale_tag.text)
            result["基金规模"] = scale_match.group(1) if scale_match else ""
    
    except (IndexError, AttributeError) as e:
        tqdm.write(f"解析基本信息时发生异常: {str(e)}")
    
    return result

def parse_performance_data(code: str) -> Dict[str, str]:
    """解析基金业绩数据"""
    url = f"https://fundf10.eastmoney.com/FundArchivesDatas.aspx?type=jdzf&code={code}"
    try:
        response = safe_request(url)
        html_content = re.findall(
            r"content:\s*\"(.*?)\"\s*};", 
            response.text, 
            re.DOTALL
        )[0]
        soup = BeautifulSoup(html_content, "html.parser")
        
        performance = {}
        for i, period in enumerate(["今年来", "近1周", "近1月", "近3月", 
                                   "近6月", "近1年", "近2年", "近3年", "近5年"], 1):
            try:
                performance[period] = soup.div.contents[i].contents[1].text
            except (IndexError, AttributeError):
                performance[period] = ""
        return performance
    except Exception as e:
        tqdm.write(f"获取业绩数据失败: {code}, 错误: {str(e)}")
        return {}

def parse_fee_info(code: str) -> Dict[str, str]:
    """解析费率信息"""
    url = f"https://fundf10.eastmoney.com/jbgk_{code}.html"
    try:
        response = safe_request(url)
        soup = BeautifulSoup(response.text, "html.parser")
        
        fee_info = {
            "管理费率": "",
            "托管费率": ""
        }
        # 管理费
        manage_tag = soup.find("th", string="管理费率")
        if manage_tag:
            fee_info["管理费率"] = manage_tag.find_next("td").text.split("%")[0] + "%"
        
        # 托管费
        custodian_tag = soup.find("th", string="托管费率")
        if custodian_tag:
            fee_info["托管费率"] = custodian_tag.find_next("td").text.split("%")[0] + "%"
        
        return fee_info
    except Exception as e:
        tqdm.write(f"获取费率信息失败: {code}, 错误: {str(e)}")
        return {}

def process_fund_data(code: str, index_data_written: bool) -> tuple:
    """处理单只基金数据"""
    try:
        # 获取基本信息
        fund_info = ak.fund_info_index_em(symbol="沪深指数", indicator="被动指数型")
        fund_data = fund_info[fund_info["基金代码"] == code].iloc[0].to_dict()
        
        # 获取网页数据
        response = safe_request(f"https://fund.eastmoney.com/{code}.html")
        soup = BeautifulSoup(response.text, "html.parser")
        
        # 组装数据
        return {
            "基金代码": code,
            "基金名称": fund_data.get("基金名称", ""),
            **parse_fund_basic_info(soup),
            **parse_performance_data(code),
            **parse_fee_info(code)
        }, index_data_written
    
    except Exception as e:
        tqdm.write(f"处理基金 {code} 时发生错误: {str(e)}")
        return None, index_data_written

def write_to_excel(ws: Worksheet, data: List[Dict], averages: Dict):
    """将处理后的数据写入Excel"""
    # 设置基础样式
    alignment = Alignment(horizontal="center", vertical="center")
    colors = CONFIG["colors"]
    
    # 写入基金数据
    for col_idx, fund_data in enumerate(data, start=3):
        col_letter = chr(65 + col_idx)
        
        # 写入基础信息
        ws[f"{col_letter}1"] = fund_data["基金代码"]
        ws[f"{col_letter}2"] = fund_data["基金名称"]
        ws[f"{col_letter}3"] = fund_data["成立日期"]
        
        # 写入业绩数据
        for row_idx, key in enumerate(CONFIG["headers"][3:12], start=4):
            cell = ws[f"{col_letter}{row_idx}"]
            cell.value = fund_data.get(key, "")
            cell.alignment = alignment
            
            # 高亮超过平均值的指标
            if key in averages and "%" in str(cell.value):
                current_value = float(cell.value.split("%")[0])
                if current_value > averages[key]:
                    cell.fill = PatternFill(
                        fill_type="solid", 
                        fgColor=colors["performance"]
                    )

def main():
    # 初始化工作环境
    os.chdir(BASE_DIR)
    wb = initialize_workbook()
    ws = wb.active
    ws.title = "基金对比"
    ws["A1"] = datetime.now().strftime("%Y-%m-%d")

    # 读取基金代码
    fund_codes = read_fund_codes()
    
    # 处理基金数据
    all_data = []
    averages = {"近1年": 0, "近2年": 0, "近3年": 0, "近5年": 0}
    index_data_written = False

    with tqdm(fund_codes, desc="处理基金数据", unit="支", ncols=100) as pbar:
        for code in pbar:
            pbar.set_postfix_str(code)
            fund_data, index_data_written = process_fund_data(code, index_data_written)
            if fund_data:
                all_data.append(fund_data)
                # 累计计算平均值
                for period in averages:
                    if fund_data[period]:
                        averages[period] += float(fund_data[period].split("%")[0])
    
    # 计算平均值
    fund_count = len(all_data)
    for period in averages:
        averages[period] /= fund_count if fund_count else 1
    
    # 写入Excel
    write_to_excel(ws, all_data, averages)
    
    # 保存文件
    wb.save(CONFIG["output_file"])
    print("处理完成！")

if __name__ == "__main__":
    main()